from openpyxl import Workbook, load_workbook
from textblob import TextBlob


class Excel:
    def __init__(self, fPath):
        self.filepath = fPath
        self.wb = Workbook()
        self.comment = {}

    def fileParser(self):
        # load workbook
        self.wb = load_workbook(self.filepath)
        # select specific worksheet
        sheet = self.wb['tbl1']
        # get max row count
        max_row = sheet.max_row
        # iterate through all rows in first column, skipping first row
        for r in range(2, max_row + 1):
            key = sheet.cell(row=r, column=1)
            cmmnt = sheet.cell(row=r, column=4)
            self.comment[key.value] = cmmnt.value
        self.wb.close()

    def sentiment(self):
        for key, value in self.comment.items():
            testimonial = TextBlob(str(value))
            self.comment[
                key] = testimonial.sentiment.polarity, \
                       testimonial.sentiment.subjectivity

    def fileWriter(self):
        # load workbook
        self.wb = load_workbook(self.filepath)
        # select specific worksheet
        sheet = self.wb['tbl2']
        # iterate through all rows in first column, skipping first row
        r = 1
        for key in self.comment.keys():
            c = 1
            r += 1
            sheet.cell(row=r, column=1).value = key
            for val in self.comment[key]:
                c += 1
                sheet.cell(row=r, column=c).value = val
        self.wb.save(self.filepath)
        self.wb.close()
